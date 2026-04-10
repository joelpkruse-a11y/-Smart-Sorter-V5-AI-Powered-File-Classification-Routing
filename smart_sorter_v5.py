import os
import sys
import time
import json
import shutil
import traceback
from datetime import datetime
import re

# Third‑party libraries
import fitz  # PyMuPDF
import docx
import google.generativeai as genai
from PIL import Image
import numpy as np

# Optional dependencies
try:
    import cv2
    HAS_CV2 = True
except Exception:
    HAS_CV2 = False

# Local modules (kept, but imports reorganized)
from ai_classifier import (
    classify_document_smart,
    generate_filename,
)
from smart_mode_v2 import smart_mode_v2
from filename_router import generate_final_filename, route_file
from v3_debug_dashboard import add_event
from metadata_enhancer import enhance_metadata

print("[DEBUG] Loaded smart_sorter_v5 (V5.1, unified AI pipeline) from:", __file__)

# ---------------------------------------------------------
# Config loading
# ---------------------------------------------------------
def load_config():
    """
    Load config.json and inject GEMINI_API_KEY from environment.
    """
    with open("config.json", "r", encoding="utf-8") as f:
        config = json.load(f)

    # Inject API key from environment
    ai_cfg = config.get("ai") or {}
    ai_cfg["api_key"] = os.getenv("GEMINI_API_KEY")
    config["ai"] = ai_cfg

    return config

# ---------------------------------------------------------
# Gemini scaffolding
# ---------------------------------------------------------
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")
GEMINI_MODEL = os.getenv("GEMINI_MODEL", "gemini-2.5-pro")


def _init_gemini():
    """
    Initialize Gemini SDK with API key from environment.
    Raises RuntimeError if key is missing.
    """
    if not GEMINI_API_KEY:
        raise RuntimeError("GEMINI_API_KEY is not set")
    genai.configure(api_key=GEMINI_API_KEY)

# ---------------------------------------------------------
# Logging
# ---------------------------------------------------------
COLOR_MAP = {
    "success": "\033[92m",
    "error": "\033[91m",
    "warn": "\033[93m",
    "diag": "\033[96m",
    "info": "\033[97m",
    "reset": "\033[0m",
}


def log(msg: str, level: str = "info"):
    """
    Color‑coded console logger, Render‑safe.
    """
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    color = COLOR_MAP.get(level, COLOR_MAP["info"])
    reset = COLOR_MAP["reset"]
    print(f"{color}[{level.upper()}] {ts} {msg}{reset}")

# ---------------------------------------------------------
# TEXT EXTRACTION
# ---------------------------------------------------------
def clean_extracted_text(text: str) -> str:
    lines = text.splitlines()
    cleaned = [line.strip() for line in lines if line.strip()]
    return "\n".join(cleaned)


def extract_text_generic(path: str) -> str:
    """
    Generic text extraction for txt, md, log, csv, docx, pdf, xlsx.
    Returns a cleaned string (may be empty on failure).
    """
    ext = os.path.splitext(path)[1].lower()

    # Plain text‑like
    if ext in [".txt", ".md", ".log", ".csv"]:
        try:
            with open(path, "r", encoding="utf-8", errors="ignore") as f:
                return f.read()
        except Exception:
            return ""

    # DOCX
    if ext == ".docx":
        try:
            doc = docx.Document(path)
            parts = [p.text.strip() for p in doc.paragraphs if p.text.strip()]
            return clean_extracted_text("\n".join(parts))
        except Exception:
            return ""

    # PDF
    if ext == ".pdf":
        try:
            doc = fitz.open(path)
            texts = [page.get_text("text") for page in doc]
            doc.close()
            return clean_extracted_text("\n".join(texts))
        except Exception:
            return ""

    # XLSX
    if ext == ".xlsx":
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, data_only=True)
            parts = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                parts.append(f"### SHEET: {sheet} ###")
                for row in ws.iter_rows(values_only=True):
                    vals = [str(v).strip() for v in row if v is not None]
                    if vals:
                        parts.append(" | ".join(vals))
            return clean_extracted_text("\n".join(parts))
        except Exception:
            return ""

    # Fallback
    return ""

# ---------------------------------------------------------
# IMAGE → CLEAN PDF
# ---------------------------------------------------------
def convert_image_to_clean_pdf(image_path: str, log_fn=log) -> str:
    """
    Convert an image to a cleaned, high‑contrast PDF for better OCR/vision.
    Returns the new PDF path, or the original image path on failure.
    """
    base, _ = os.path.splitext(image_path)
    pdf_path = base + "_cleaned.pdf"

    try:
        img = Image.open(image_path).convert("RGB")
        img_np = np.array(img)

        if img_np.ndim == 3 and HAS_CV2:
            gray = cv2.cvtColor(img_np, cv2.COLOR_RGB2GRAY)
            bw = cv2.adaptiveThreshold(
                gray,
                255,
                cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                cv2.THRESH_BINARY,
                35,
                10,
            )
            Image.fromarray(bw).save(pdf_path, "PDF", resolution=300)
        else:
            img.convert("L").save(pdf_path, "PDF", resolution=300)

        log_fn(f"[PDF] Converted to cleaned PDF: {pdf_path}", "diag")
        return pdf_path

    except Exception as e:
        log_fn(f"[PDF] Conversion failed: {e}", "warn")
        return image_path

# ---------------------------------------------------------
# GEMINI: UNIFIED CLASSIFICATION + SUMMARY
# ---------------------------------------------------------
def gemini_process_document(
    *,
    path: str,
    text: str,
    filename: str,
    config: dict,
    log_fn=log,
) -> dict:
    """
    Unified Gemini call:
      - Classify into a category
      - Provide confidence
      - Provide reasoning
      - Suggest semantic filename
      - Provide a concise summary
      - Return STRICT JSON

    Returns a dict with at least:
      - text
      - category
      - confidence
      - metadata
      - tables
      - filename
      - reasoning
      - summary
    """
    _init_gemini()

    ai_cfg = (config.get("ai") or {})
    filename_style = ai_cfg.get("filename_style", "semantic-kebab")

    classification_cfg = (config.get("classification") or {})
    categories = classification_cfg.get("categories", [])
    category_names = [c.get("name", "other") for c in categories]

    # Truncate text for prompt safety
    prompt_text = (text or "")[:4000]

    prompt = f"""
You are Smart Sorter V5.1 running on {GEMINI_MODEL}.

Your job:
1. Classify the document into ONE of these categories (best fit only):
   {category_names}
2. Provide a confidence score between 0 and 1 (float).
3. Provide a short reasoning string explaining the classification.
4. Suggest a semantic filename using style "{filename_style}" (no extension).
5. Provide a concise, user‑friendly summary of the document (2–6 sentences).
6. Return STRICT JSON only. No extra text, no markdown, no commentary.

Input filename: {filename}

Document text (truncated):
\"\"\"{prompt_text}\"\"\"

JSON response shape (STRICT):
{{
  "text": "string",          // cleaned or representative text (may reuse input)
  "category": "string",      // one of {category_names}
  "confidence": 0.0,         // float between 0 and 1
  "metadata": {{}},          // optional structured metadata
  "tables": [],              // optional table structures
  "filename": "string",      // semantic filename (no extension)
  "reasoning": "string",     // short explanation of classification
  "summary": "string"        // concise user‑friendly summary
}}
"""

    model = genai.GenerativeModel(GEMINI_MODEL)
    resp = model.generate_content(prompt)

    try:
        data = json.loads(resp.text)
    except Exception as e:
        log_fn(f"[AI] Gemini JSON parse failed: {e}", "error")
        return {}

    # Defaults / normalization
    data.setdefault("text", text or "")
    data.setdefault("category", "other")
    data.setdefault("confidence", 0.0)
    data.setdefault("metadata", {})
    data.setdefault("tables", [])
    data.setdefault("filename", filename)
    data.setdefault("reasoning", "")
    data.setdefault("summary", "")

    # Type safety
    try:
        data["confidence"] = float(data.get("confidence", 0.0))
    except Exception:
        data["confidence"] = 0.0

    return data

    # ---------------------------------------------------------
# INTERNAL PIPELINE (V5.1 unified AI)
# ---------------------------------------------------------
def _classify_and_route_internal(path: str, config: dict) -> dict:
    """
    Internal pipeline:
      - Validate file
      - Determine type (photo, video, document)
      - Extract text (for documents)
      - Call Gemini once for classification + summary
      - Apply Smart Mode V2
      - Generate final filename
      - Route file
      - Emit dashboard event
      - Return a rich result dict for the web dashboard
    """
    if not os.path.exists(path):
        return {"status": "Failed", "summary": "File missing"}

    original_name = os.path.basename(path)
    ext = os.path.splitext(original_name)[1].lower()

    classification_cfg = (config.get("classification") or {})
    photo_exts = [e.lower() for e in classification_cfg.get("photo_extensions", [])]
    video_exts = [e.lower() for e in classification_cfg.get("video_extensions", [])]

    # Filesystem metadata
    try:
        stat = os.stat(path)
        metadata_fs = {
            "created": datetime.fromtimestamp(stat.st_ctime).isoformat(),
            "modified": datetime.fromtimestamp(stat.st_mtime).isoformat(),
        }
    except Exception:
        metadata_fs = {}

    # Defaults
    category = "other"
    confidence = 0.0
    ai_filename = None
    text = ""
    extracted_text = ""
    metadata = dict(metadata_fs)
    reasoning = ""
    summary_text = ""
    treat_as_document = False

    classification_path = path
    storage_path = path

    # XLSX always treated as document
    if ext == ".xlsx":
        treat_as_document = True

    # PHOTO DETECTION
    if ext in photo_exts:
        # For now, keep photos as non‑document for Render safety
        treat_as_document = False
        if not treat_as_document:
            category = "photos"
            confidence = 1.0
            ai_filename = original_name

    # VIDEO DETECTION
    elif ext in video_exts:
        category = "videos"
        confidence = 1.0
        ai_filename = original_name

    # DOCUMENT‑PHOTO → PDF (if we ever flip treat_as_document=True for photos)
    if ext in photo_exts and treat_as_document:
        storage_path = convert_image_to_clean_pdf(path, log)
        classification_path = storage_path

    # NON‑PHOTO DOCUMENT → extract text
    if ext not in photo_exts:
        extracted_text = extract_text_generic(classification_path)

    # GEMINI CLASSIFICATION + SUMMARY (documents and also as a fallback for others)
    gemini_result = gemini_process_document(
        path=classification_path,
        text=extracted_text,
        filename=original_name,
        config=config,
        log_fn=log,
    )

    if gemini_result:
        text = gemini_result.get("text", "") or extracted_text
        category = gemini_result.get("category", category or "other")
        confidence = float(gemini_result.get("confidence", confidence or 0.0))
        metadata_ai = gemini_result.get("metadata") or {}
        ai_filename = gemini_result.get("filename") or ai_filename or original_name
        reasoning = gemini_result.get("reasoning", "") or reasoning
        summary_text = gemini_result.get("summary", "") or summary_text

        # Metadata enhancement
        metadata = enhance_metadata(
            text=text,
            metadata_ai=metadata_ai,
            metadata_vision={},
            metadata_fs=metadata_fs,
        )

    # FALLBACK FILENAME if Gemini didn't provide one
    if not ai_filename:
        ai_filename = generate_filename(
            text=text,
            category=category,
            original_filename=original_name,
            metadata=metadata,
        )

    # SMART MODE V2
    result_for_smart = {
        "category": category,
        "confidence": confidence,
        "metadata": metadata,
        "text": text,
    }
    final_category = smart_mode_v2(result_for_smart, log) or category

    # ROUTING
    result_for_router = {
        "category": final_category,
        "confidence": confidence,
        "metadata": metadata,
        "text": text,
        "filename": ai_filename,
    }

    final_filename = generate_final_filename(result_for_router, storage_path, log)
    destination = route_file(final_category, final_filename, config, log)

    # MOVE FILE
    try:
        shutil.move(storage_path, destination)
    except Exception as e:
        log(f"[MOVE] Failed: {e}", "error")

    # DASHBOARD EVENT
    add_event(
        {
            "original": original_name,
            "category": final_category,
            "confidence": confidence,
            "final_filename": final_filename,
            "metadata": metadata,
            "reasoning": reasoning,
            "text": text,
        }
    )

    # Final summary fallback chain
    final_summary = summary_text or reasoning or text[:200]

    return {
        "status": "Completed",
        "category": final_category,
        "confidence": confidence,
        "summary": final_summary,
        "final_filename": final_filename,
        "extracted_text": extracted_text,
        "ocr_text": None,
        "text": text,
    }
# ---------------------------------------------------------
# RENDER‑SAFE ENTRYPOINT FOR FLASK DASHBOARD
# ---------------------------------------------------------
def process_file_for_web(path: str, config: dict) -> dict:
    """
    Unified entrypoint for the Flask dashboard.

    Produces a dict with at least:
      - status
      - category
      - confidence
      - summary
      - final_filename
      - extracted_text
      - ocr_text
      - text
    """
    try:
        result = _classify_and_route_internal(path, config)
        if not result:
            return {"status": "Failed", "summary": "Empty result from pipeline"}
        return result

    except Exception as e:
        log(f"[WEB] Error: {e}", "error")
        traceback.print_exc()
        return {"status": "Failed", "summary": str(e)}



