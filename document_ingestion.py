import sqlite3
import hashlib
import json
from pathlib import Path
from datetime import datetime

from PIL import Image, ImageDraw, ImageFont
from utils.ocr_utils import extract_text_for_ai

try:
    from pdf2image import convert_from_path
except ImportError:
    convert_from_path = None

try:
    from docx import Document
except ImportError:
    Document = None

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


# ------------------------------------------------------------
# CONFIG PATH RESOLUTION & DATABASE CONNECTION
# ------------------------------------------------------------
def _get_ingestion_paths(config: dict):
    """Dynamically resolves paths from config with safe fallbacks."""
    config = config or {}
    
    # Base system path (fallback to C:/SmartInbox if not in config)
    sys_root = Path(config.get("system_root", "C:/SmartInbox"))
    db_path = sys_root / "faces.db"
    
    # Thumbnails
    sorted_root = Path(config.get("destinations", {}).get("sorted_root", "C:/SmartInbox/Photos"))
    thumb_root = sorted_root / "Thumbs"
    
    thumb_root.mkdir(parents=True, exist_ok=True)
    
    return db_path, thumb_root

def _get_db_connection(db_path: Path):
    """Returns a highly concurrent SQLite connection."""
    # 15-second timeout handles write-contention gracefully
    conn = sqlite3.connect(db_path, timeout=15.0)
    conn.row_factory = sqlite3.Row
    # Write-Ahead Logging allows simultaneous readers and writers
    conn.execute("PRAGMA journal_mode=WAL;")
    conn.execute("PRAGMA synchronous=NORMAL;") 
    return conn

# ------------------------------------------------------------
# SHA256
# ------------------------------------------------------------
def compute_sha256(path: Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()

# ------------------------------------------------------------
# PREVIEW GENERATORS
# ------------------------------------------------------------
def _render_text_preview(text: str, sha: str, thumb_root: Path):
    out_path = thumb_root / f"{sha}.jpg"
    if out_path.exists():
        return out_path

    img = Image.new("RGB", (800, 1000), color=(255, 255, 255))
    draw = ImageDraw.Draw(img)
    try:
        font = ImageFont.truetype("arial.ttf", 20)
    except Exception:
        font = ImageFont.load_default()

    draw.text((20, 20), text[:2000], fill=(0, 0, 0), font=font)
    img.thumbnail((400, 400))
    img.save(out_path, "JPEG", quality=80)
    return out_path

def generate_pdf_preview(path: Path, sha: str, thumb_root: Path):
    out_path = thumb_root / f"{sha}.jpg"
    if out_path.exists():
        return out_path

    if convert_from_path:
        try:
            pages = convert_from_path(path, dpi=72, last_page=1)
            if pages:
                img = pages[0]
                img.thumbnail((400, 400))
                img.save(out_path, "JPEG", quality=80)
                return out_path
        except Exception:
            pass
    return None

def generate_docx_preview(path: Path, sha: str, thumb_root: Path):
    if not Document:
        return None
    try:
        doc = Document(path)
        text = "\n".join([p.text for p in doc.paragraphs[:20]])
        return _render_text_preview(text, sha, thumb_root)
    except Exception:
        return None

def generate_xlsx_preview(path: Path, sha: str, thumb_root: Path):
    if not load_workbook:
        return None
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        lines = []
        for row in ws.iter_rows(min_row=1, max_row=20, values_only=True):
            lines.append(" | ".join([str(v) if v else "" for v in row]))
        text = "\n".join(lines)
        return _render_text_preview(text, sha, thumb_root)
    except Exception:
        return None

def generate_icon_thumbnail(doc_type: str, sha: str, thumb_root: Path):
    out_path = thumb_root / f"{sha}.jpg"
    if out_path.exists():
        return out_path

    img = Image.new("RGB", (400, 400), color=(240, 240, 240))
    draw = ImageDraw.Draw(img)
    try:
        font = ImageFont.truetype("arial.ttf", 60)
    except Exception:
        font = ImageFont.load_default()

    draw.text((100, 160), doc_type, fill=(100, 100, 100), font=font)
    img.save(out_path, "JPEG", quality=80)
    return out_path

# ------------------------------------------------------------
# DATABASE INGESTION
# ------------------------------------------------------------
def _init_db(conn):
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS photos (
            sha256 TEXT PRIMARY KEY,
            filename TEXT,
            original_path TEXT,
            thumb_path TEXT,
            face_thumb_path TEXT,
            face_count INTEGER,
            tags TEXT,
            tag_scores TEXT,
            face_data TEXT,
            date_taken TEXT,
            gps_lat REAL,
            gps_lon REAL,
            created_at TEXT,
            camera_model TEXT,
            ocr_text TEXT,
            category TEXT,
            doc_type TEXT,
            doc_issuer TEXT,
            is_document INTEGER DEFAULT 0
        )
        """
    )
    conn.commit()

def ingest_document(path: Path, config: dict = None):
    """
    Ingests DOCX, PDF, TXT, XLSX, etc. into the DB.
    Now uses robust WAL mode and dynamic paths.
    """
    db_path, thumb_root = _get_ingestion_paths(config)
    
    sha = compute_sha256(path)
    filename = path.name

    suffix = path.suffix.lower()
    if suffix == ".pdf":
        doc_type = "PDF"
    elif suffix == ".docx":
        doc_type = "DOCX"
    elif suffix == ".xlsx":
        doc_type = "XLSX"
    elif suffix in [".txt", ".md"]:
        doc_type = "TEXT"
    else:
        doc_type = "FILE"

    # Generate preview/thumbnail
    if suffix == ".pdf":
        thumb_path = generate_pdf_preview(path, sha, thumb_root)
    elif suffix == ".docx":
        thumb_path = generate_docx_preview(path, sha, thumb_root)
    elif suffix == ".xlsx":
        thumb_path = generate_xlsx_preview(path, sha, thumb_root)
    elif suffix in [".txt", ".md"]:
        try:
            text_raw = path.read_text(encoding="utf-8", errors="ignore")
        except Exception:
            text_raw = ""
        thumb_path = _render_text_preview(text_raw[:4000], sha, thumb_root)
    else:
        thumb_path = generate_icon_thumbnail(doc_type, sha, thumb_root)

    # Extract OCR/Text for search
    try:
        ocr_text = extract_text_for_ai(path)
    except Exception:
        ocr_text = ""

    # Write to database safely
    conn = _get_db_connection(db_path)
    _init_db(conn)

    conn.execute(
        """
        INSERT OR REPLACE INTO photos (
            sha256, filename, original_path, thumb_path,
            is_document, doc_type, ocr_text, created_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            sha,
            filename,
            str(path),
            str(thumb_path) if thumb_path else None,
            1,
            doc_type,
            ocr_text,
            datetime.now().isoformat(),
        )
    )

    conn.commit()
    conn.close()